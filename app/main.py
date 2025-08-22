# -*- coding: utf-8 -*-
import pandas as pd
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox
from datetime import datetime
from typing import Optional
from openpyxl.styles import PatternFill
from openpyxl import load_workbook

YELLOW_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")

# ---------- ВСПОМОГАТЕЛЬНОЕ ----------
def _norm_cols(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    return df

def _find_col(cols: list[str], name: str) -> Optional[int]:
    low = [c.casefold() for c in cols]
    return low.index(name.casefold()) if name.casefold() in low else None

def _get_col_name_case_insensitive(cols: list[str], target: str) -> Optional[str]:
    idx = _find_col(cols, target)
    return cols[idx] if idx is not None else None

# ---------- ЛОГИКА ----------
def load_base(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path)
    df = _norm_cols(df)
    if "MID" not in df.columns:
        raise ValueError("В файле базы не найден столбец 'MID'.")
    if "TID" not in df.columns:
        df["TID"] = ""
    df = df[["MID", "TID"]].copy()
    df["MID"] = df["MID"].astype(str).str.strip()
    df["TID"] = df["TID"].astype(str).str.strip()
    # агрегируем дубль MID → уникальные TID через запятую
    base_map = (
        df.groupby("MID", as_index=False)
          .agg({"TID": lambda s: ", ".join(sorted({x for x in s if x}))})
    )
    return base_map

def load_list(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path)
    df = _norm_cols(df)
    # обязательна только колонка АЗС (в любом регистре)
    azs_name = _get_col_name_case_insensitive(list(df.columns), "АЗС")
    if not azs_name:
        raise ValueError("В файле списка отсутствует столбец 'АЗС'.")
    df[azs_name] = df[azs_name].astype(str).str.strip()
    return df

def insert_new_terminal_column(df: pd.DataFrame, new_col_name="новый терминал") -> pd.DataFrame:
    cols = list(df.columns)
    idx_terminal = _find_col(cols, "терминал")
    idx_addr = _find_col(cols, "Адрес")
    if idx_terminal is not None:
        insert_at = idx_terminal + 1
    elif idx_addr is not None:
        insert_at = idx_addr
    else:
        insert_at = len(cols)
    out = df.copy()
    if new_col_name not in out.columns:
        out.insert(insert_at, new_col_name, "")
    else:
        # если вдруг уже есть — просто переставим
        cols.remove(new_col_name)
        cols.insert(insert_at, new_col_name)
        out = out.reindex(columns=cols)
    return out

def enrich_and_style(list_df: pd.DataFrame, base_df: pd.DataFrame, partner_name: str, out_dir: Path) -> Path:
    base_map = dict(zip(base_df["MID"], base_df["TID"]))
    out_df = insert_new_terminal_column(list_df, "новый терминал")

    azs_col = _get_col_name_case_insensitive(list(out_df.columns), "АЗС")
    if not azs_col:
        raise ValueError("Не удалось найти колонку 'АЗС' после вставки столбца.")

    out_df.loc[:, azs_col] = out_df[azs_col].astype(str).str.strip()
    mask = out_df[azs_col].isin(base_map.keys())
    out_df.loc[mask, "новый терминал"] = out_df.loc[mask, azs_col].map(base_map)

    out_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = (partner_name or "Partner").strip().replace(" ", "_")
    out_path = out_dir / f"PartnerList_Enriched_{safe_name}_{stamp}.xlsx"

    # Сохраняем
    out_df.to_excel(out_path, index=False, sheet_name="Лист партнера")

    # Подкраска «АЗС» жёлтым для совпадений
    wb = load_workbook(out_path)
    ws = wb.active
    header_map = {cell.value: cell.column for cell in ws[1] if cell.value is not None}
    azs_idx = header_map.get(azs_col)
    if azs_idx:
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=azs_idx).value
            if v is not None and str(v).strip() in base_map:
                ws.cell(row=r, column=azs_idx).fill = YELLOW_FILL
    wb.save(out_path)

    return out_path

# ---------- GUI ----------
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Дополнение списка TID из базы (MID ↔ АЗС)")
        self.geometry("700x380")
        self.minsize(660, 360)

        self.base_path: Optional[Path] = None
        self.list_path: Optional[Path] = None

        bar = tk.Frame(self); bar.pack(fill="x", padx=12, pady=12)
        self.btn_base = tk.Button(bar, text="📘 База терминалов", height=2, command=self.pick_base); self.btn_base.grid(row=0, column=0, padx=6, pady=6, sticky="ew")
        self.btn_list = tk.Button(bar, text="📄 Список для партнёра", height=2, command=self.pick_list); self.btn_list.grid(row=0, column=1, padx=6, pady=6, sticky="ew")
        bar.columnconfigure(0, weight=1); bar.columnconfigure(1, weight=1)

        paths = tk.Frame(self); paths.pack(fill="x", padx=12)
        tk.Label(paths, text="База:").grid(row=0, column=0, sticky="w")
        tk.Label(paths, text="Список:").grid(row=1, column=0, sticky="w")
        self.lbl_base = tk.Label(paths, text="— не выбран —", anchor="w", fg="#555"); self.lbl_base.grid(row=0, column=1, sticky="w", padx=6)
        self.lbl_list = tk.Label(paths, text="— не выбран —", anchor="w", fg="#555"); self.lbl_list.grid(row=1, column=1, sticky="w", padx=6)

        partner = tk.Frame(self); partner.pack(fill="x", padx=12, pady=10)
        tk.Label(partner, text="Введите название партнёра:").pack(anchor="w")
        self.partner_entry = tk.Entry(partner); self.partner_entry.pack(fill="x")

        self.btn_go = tk.Button(self, text="✨ Дополнить", height=2, state="disabled", command=self.run); self.btn_go.pack(pady=16)

        self.status = tk.StringVar(value="Готово")
        status_bar = tk.Label(self, textvariable=self.status, bd=1, relief="sunken", anchor="w"); status_bar.pack(side="bottom", fill="x")

    def pick_base(self):
        path = filedialog.askopenfilename(title="Выберите файл базы (Excel)", filetypes=[("Excel", "*.xlsx *.xls"), ("Все файлы", "*.*")])
        if path:
            self.base_path = Path(path); self.lbl_base.config(text=str(self.base_path)); self._toggle_go()

    def pick_list(self):
        path = filedialog.askopenfilename(title="Выберите файл списка (Excel)", filetypes=[("Excel", "*.xlsx *.xls"), ("Все файлы", "*.*")])
        if path:
            self.list_path = Path(path); self.lbl_list.config(text=str(self.list_path)); self._toggle_go()

    def _toggle_go(self):
        self.btn_go.config(state="normal" if (self.base_path and self.list_path) else "disabled")

    def run(self):
        try:
            self.status.set("Читаю базу..."); self.update_idletasks()
            base_df = load_base(self.base_path)

            self.status.set("Читаю список..."); self.update_idletasks()
            list_df = load_list(self.list_path)

            self.status.set("Обрабатываю..."); self.update_idletasks()
            partner_name = self.partner_entry.get().strip() or "Partner"
            desktop = Path.home() / "Desktop"
            out_path = enrich_and_style(list_df, base_df, partner_name, desktop)

            self.status.set("Готово")
            messagebox.showinfo("Готово", f"Файл создан:\n{out_path}")
        except Exception as e:
            self.status.set("Ошибка")
            messagebox.showerror("Ошибка", f"{e}")

if __name__ == "__main__":
    App().mainloop()